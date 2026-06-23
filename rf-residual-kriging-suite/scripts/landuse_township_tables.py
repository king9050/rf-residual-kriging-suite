import os
import re
import math
import uuid
import datetime as dt

import arcpy
from arcpy.sa import Con, Int, SetNull, ExtractByMask, ZonalStatisticsAsTable, TabulateArea

import numpy as np
import openpyxl


def _norm(s):
    if s is None:
        return ""
    s = str(s)
    s = s.replace("\u3000", " ")
    s = s.replace(" ", "")
    s = s.strip()
    return s


def _safe_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        try:
            x = float(v)
            if math.isnan(x) or math.isinf(x):
                return None
            return x
        except Exception:
            return None
    try:
        s = str(v).strip()
        if s == "":
            return None
        x = float(s)
        if math.isnan(x) or math.isinf(x):
            return None
        return x
    except Exception:
        return None


def _ensure_dir(p):
    os.makedirs(p, exist_ok=True)
    return p


def _create_gdb(folder, name):
    gdb = os.path.join(folder, name)
    if not arcpy.Exists(gdb):
        arcpy.management.CreateFileGDB(folder, name)
    return gdb


def _choose_sheet(wb):
    # 优先选择包含特定关键词的表
    preferred_keywords = ["数据清洗", "修正后", "clean", "final"]
    for name in wb.sheetnames:
        if any(keyword in name for keyword in preferred_keywords):
            return name
    
    # 如果没有找到优先表，再选择行数最多的
    best = None
    best_count = -1
    for name in wb.sheetnames:
        ws = wb[name]
        count = 0
        for r in ws.iter_rows(min_row=2, values_only=True):
            if r and any(v is not None and str(v).strip() != "" for v in r):
                count += 1
            if count >= 5000:
                break
        if count > best_count:
            best = name
            best_count = count
    return best


def _parse_range_spec(spec):
    if spec is None:
        return []
    s = str(spec).strip()
    if s == "":
        return []
    s = s.replace("≤", "<=").replace("≥", ">=")
    s = s.replace("，", ",")
    tokens = []
    tokens.extend(re.findall(r"[\(\[]\s*[-\d.]+\s*,\s*[-\d.]+\s*[\)\]]", s))
    tokens.extend(re.findall(r"(?:<=|>=|<|>)\s*[-\d.]+", s))
    out = []
    for t in tokens:
        t = t.replace(" ", "")
        if t.startswith(("(", "[")) and t.endswith((")", "]")):
            left_inc = t[0] == "["
            right_inc = t[-1] == "]"
            body = t[1:-1]
            a, b = body.split(",", 1)
            lo = _safe_float(a)
            hi = _safe_float(b)
            if lo is None or hi is None:
                continue
            out.append((lo, hi, left_inc, right_inc))
        else:
            m = re.match(r"(<=|>=|<|>)\s*([-\d.]+)", t)
            if not m:
                continue
            op = m.group(1)
            x = _safe_float(m.group(2))
            if x is None:
                continue
            if op == ">":
                out.append((x, None, False, False))
            elif op == ">=":
                out.append((x, None, True, False))
            elif op == "<":
                out.append((None, x, False, False))
            elif op == "<=":
                out.append((None, x, False, True))
    return out


def _segs_to_raster_cond(r, segs):
    cond = None
    for lo, hi, lo_inc, hi_inc in segs:
        c = None
        if lo is not None:
            c = (r >= lo) if lo_inc else (r > lo)
        if hi is not None:
            c2 = (r <= hi) if hi_inc else (r < hi)
            c = c2 if c is None else (c & c2)
        if c is None:
            continue
        cond = c if cond is None else (cond | c)
    return cond


def _read_standards(standard_xlsx):
    wb = openpyxl.load_workbook(standard_xlsx, data_only=False)
    ws = wb[wb.sheetnames[0]]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    attr_cols = {}
    for c in range(3, ws.max_column + 1):
        v = headers[c - 1]
        if v is None:
            continue
        attr = _norm(v)
        if attr != "":
            attr_cols[attr] = c

    standards = {}
    display = {}
    for cls_row in range(2, 7):
        cls = ws.cell(row=cls_row, column=2).value
        cls = int(cls) if cls is not None else None
        if cls is None:
            continue
        for attr, col in attr_cols.items():
            v = ws.cell(row=cls_row, column=col).value
            if v is None or str(v).strip() == "":
                continue
            standards.setdefault(attr, {}).setdefault(cls, []).extend(_parse_range_spec(v))
            display.setdefault(attr, {})[cls] = str(v).strip()
    return standards, display


def _read_samples(sample_xlsx):
    wb = openpyxl.load_workbook(sample_xlsx, data_only=True, read_only=True)
    sheet = _choose_sheet(wb)
    ws = wb[sheet]
    head = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {_norm(h): i for i, h in enumerate(head) if h is not None and _norm(h) != ""}
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r:
            continue
        if "longitude" in idx and "latitude" in idx:
            lon = _safe_float(r[idx["longitude"]] if idx["longitude"] < len(r) else None)
            lat = _safe_float(r[idx["latitude"]] if idx["latitude"] < len(r) else None)
            if lon is None or lat is None:
                continue
        rows.append(r)
    return {"sheet": sheet, "header_idx": idx, "rows": rows}


def _published_areas_mu():
    paddy = 296564
    irrigated = 5
    dryland = 30033
    orchard = 9834
    tea = 55270
    rubber = 0
    other_garden = 17576
    forest = 3484725
    grass = 18257
    total = 4277101
    cropland = paddy + irrigated + dryland
    garden = orchard + tea + rubber + other_garden
    other = 364836
    return {
        "行政区总面积": total,
        "耕地": cropland,
        "耕地|水田": paddy,
        "耕地|水浇地": irrigated,
        "耕地|旱地": dryland,
        "园地": garden,
        "园地|果园": orchard,
        "园地|茶园": tea,
        "园地|橡胶园": rubber,
        "园地|其他园地": other_garden,
        "林地": forest,
        "草地": grass,
        "其他": other,
    }


def _infer_zone_from_landuse(dlmc, l1):
    dlmc = _norm(dlmc)
    l1 = _norm(l1)
    if "水田" in dlmc:
        return "耕地|水田", "耕地"
    if "水浇地" in dlmc or "水浇" in dlmc:
        return "耕地|水浇地", "耕地"
    if "旱地" in dlmc:
        return "耕地|旱地", "耕地"
    if "果园" in dlmc:
        return "园地|果园", "园地"
    if "茶园" in dlmc:
        return "园地|茶园", "园地"
    if "橡胶园" in dlmc or "橡胶" in dlmc:
        return "园地|橡胶园", "园地"
    if "园地" in dlmc:
        return "园地|其他园地", "园地"
    if l1 in ("林地", "草地", "耕地", "园地"):
        return l1, l1
    return "其他", "其他"


def _add_landuse_zone_fields(landuse_fc, zone_field="ZONE", l1_field="L1GROUP"):
    fields = [f.name for f in arcpy.ListFields(landuse_fc)]
    if zone_field not in fields:
        arcpy.management.AddField(landuse_fc, zone_field, "TEXT", field_length=40)
    if l1_field not in fields:
        arcpy.management.AddField(landuse_fc, l1_field, "TEXT", field_length=20)

    with arcpy.da.UpdateCursor(landuse_fc, ["DLMC", "一级地", zone_field, l1_field]) as cur:
        for dlmc, l1, z, g in cur:
            z2, g2 = _infer_zone_from_landuse(dlmc, l1)
            cur.updateRow([dlmc, l1, z2, g2])


def _build_zone_rasters(work_gdb, landuse_fc, snap_raster):
    zone_ras = os.path.join(work_gdb, "zone10")
    l1_ras = os.path.join(work_gdb, "l1_5")

    if arcpy.Exists(zone_ras) and arcpy.Exists(l1_ras):
        return zone_ras, l1_ras

    arcpy.env.snapRaster = snap_raster
    arcpy.env.cellSize = snap_raster
    arcpy.env.extent = snap_raster

    fields = [f.name for f in arcpy.ListFields(landuse_fc)]
    if "ZONE_ID" not in fields:
        arcpy.management.AddField(landuse_fc, "ZONE_ID", "SHORT")
    if "L1_ID" not in fields:
        arcpy.management.AddField(landuse_fc, "L1_ID", "SHORT")

    zone_map = {
        "耕地|水田": 1,
        "耕地|水浇地": 2,
        "耕地|旱地": 3,
        "园地|果园": 4,
        "园地|茶园": 5,
        "园地|橡胶园": 6,
        "园地|其他园地": 7,
        "林地": 8,
        "草地": 9,
        "其他": 10,
    }
    l1_map = {"耕地": 1, "园地": 2, "林地": 3, "草地": 4, "其他": 5}

    with arcpy.da.UpdateCursor(landuse_fc, ["ZONE", "L1GROUP", "ZONE_ID", "L1_ID"]) as cur:
        for z, g, zid, gid in cur:
            z = "" if z is None else str(z)
            g = "" if g is None else str(g)
            cur.updateRow([z, g, zone_map.get(z, 10), l1_map.get(g, 5)])

    if not arcpy.Exists(zone_ras):
        print("正在生成土地利用分区栅格: zone10")
        arcpy.conversion.PolygonToRaster(landuse_fc, "ZONE_ID", zone_ras, cell_assignment="CELL_CENTER")
    if not arcpy.Exists(l1_ras):
        print("正在生成一级地类栅格: l1_5")
        arcpy.conversion.PolygonToRaster(landuse_fc, "L1_ID", l1_ras, cell_assignment="CELL_CENTER")
    return zone_ras, l1_ras


def _zonal_mean_by_zone_raster(zone_raster, raster_path, out_table):
    ZonalStatisticsAsTable(zone_raster, "Value", raster_path, out_table, ignore_nodata="DATA", statistics_type="MEAN")
    d = {}
    with arcpy.da.SearchCursor(out_table, ["Value", "MEAN"]) as cur:
        for z, m in cur:
            if z is None:
                continue
            d[int(z)] = None if m is None else float(m)
    return d


def _sample_landuse_group(landuse_value):
    s = _norm(landuse_value)
    if "水田" in s:
        return "耕地|水田"
    if "水浇地" in s or "水浇" in s:
        return "耕地|水浇地"
    if "旱地" in s:
        return "耕地|旱地"
    if "果园" in s:
        return "园地|果园"
    if "茶园" in s:
        return "园地|茶园"
    if "橡胶" in s:
        return "园地|橡胶园"
    if "园地" in s:
        return "园地|其他园地"
    if "林地" in s:
        return "林地"
    if "草地" in s:
        return "草地"
    return "其他"


def _format_range(a, b, ndigits=2):
    if a is None or b is None:
        return "—"
    return f"{round(a, ndigits)}-{round(b, ndigits)}"


def _format_mean(v, ndigits=2):
    if v is None:
        return "—"
    return round(float(v), ndigits)


def _format_area_wanmu(mu, ndigits=3):
    if mu is None:
        return "—"
    return round(float(mu) / 10000.0, ndigits)


def _apply_table_style(ws, top, left, bottom, right):
    from openpyxl.styles import Border, Side, Alignment, Font

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(top, bottom + 1):
        for c in range(left, right + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.cell(row=top, column=left).font = Font(bold=True)


def _write_table1_sheet(ws, attr, unit, rows):
    ws.cell(row=1, column=1, value=f"各土壤属性在不同土地利用下的样点与预测栅格的面积对比表（{attr}）")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)

    ws.cell(row=2, column=1, value="土地利用类型")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
    ws.cell(row=2, column=3, value="样点统计")
    ws.merge_cells(start_row=2, start_column=3, end_row=2, end_column=5)
    ws.cell(row=2, column=6, value="制图统计")
    ws.merge_cells(start_row=2, start_column=6, end_row=2, end_column=7)

    ws.cell(row=3, column=1, value="一级")
    ws.cell(row=3, column=2, value="二级")
    ws.cell(row=3, column=3, value=f"均值({unit})")
    ws.cell(row=3, column=4, value=f"范围({unit})")
    ws.cell(row=3, column=5, value="数量/个")
    ws.cell(row=3, column=6, value=f"均值({unit})")
    ws.cell(row=3, column=7, value="面积/万亩")

    r0 = 4
    for i, r in enumerate(rows):
        rr = r0 + i
        ws.cell(row=rr, column=1, value=r["l1"])
        ws.cell(row=rr, column=2, value=r["l2"])
        ws.cell(row=rr, column=3, value=r["s_mean"])
        ws.cell(row=rr, column=4, value=r["s_rng"])
        ws.cell(row=rr, column=5, value=r["s_n"])
        ws.cell(row=rr, column=6, value=r["m_mean"])
        ws.cell(row=rr, column=7, value=r["m_area"])

    start = r0
    while start < r0 + len(rows):
        v = ws.cell(row=start, column=1).value
        end = start
        while end + 1 < r0 + len(rows) and ws.cell(row=end + 1, column=1).value == v:
            end += 1
        if v not in (None, "", "全县") and end > start:
            ws.merge_cells(start_row=start, start_column=1, end_row=end, end_column=1)
        start = end + 1

    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    for r in range(1, r0 + len(rows) + 2):
        ws.row_dimensions[r].height = 18
    _apply_table_style(ws, 1, 1, r0 + len(rows) - 1, 7)


def _sum_preserving_round(values, target_sum):
    values = np.asarray(values, dtype=float)
    floors = np.floor(values)
    frac = values - floors
    ints = floors.astype(int)
    diff = int(target_sum - int(np.sum(ints)))
    if diff > 0:
        order = np.argsort(-frac)
        for k in order[:diff]:
            ints[k] += 1
    elif diff < 0:
        order = np.argsort(frac)
        for k in order[: (-diff)]:
            if ints[k] > 0:
                ints[k] -= 1
    return ints


def _make_class_raster(raster_path, specs, out_path):
    r = arcpy.Raster(raster_path)
    base = SetNull(r == r, r)
    outc = base
    for cls in range(1, 6):
        segs = specs.get(cls, [])
        cond = _segs_to_raster_cond(r, segs)
        if cond is None:
            continue
        outc = Con(cond, cls, outc)
    Int(outc).save(out_path)
    return out_path


def _tabulate_area_mu(zone_fc, zone_field, class_raster_path, out_table):
    TabulateArea(zone_fc, zone_field, class_raster_path, "Value", out_table)
    fields = [f.name for f in arcpy.ListFields(out_table)]
    class_fields = [f for f in fields if f.lower().startswith("value_")]
    class_fields.sort(key=lambda x: int(x.split("_", 1)[1]))
    out = {}
    with arcpy.da.SearchCursor(out_table, [zone_field] + class_fields) as cur:
        for row in cur:
            z = row[0]
            if z is None:
                continue
            vals = []
            for v in row[1:]:
                vals.append(0.0 if v is None else float(v))
            out[str(z)] = {"class_fields": class_fields, "areas_mapunit": vals}
    return out


def _table2_write_sheet(ws, title, unit, township_names, class_labels, matrix_mu):
    ws.cell(row=1, column=1, value=title)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + len(class_labels))
    ws.cell(row=2, column=1, value="乡镇")
    ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
    ws.cell(row=2, column=2, value=f"分级（{unit}）")
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=1 + len(class_labels))
    for j, lab in enumerate(class_labels, start=2):
        ws.cell(row=3, column=j, value=lab)
    ws.cell(row=3, column=2 + len(class_labels), value="总计/亩")

    r0 = 4
    for i, name in enumerate(township_names):
        rr = r0 + i
        ws.cell(row=rr, column=1, value=name)
        row_vals = matrix_mu[i]
        for j, v in enumerate(row_vals, start=2):
            ws.cell(row=rr, column=j, value=int(v))
        ws.cell(row=rr, column=2 + len(class_labels), value=int(np.sum(row_vals)))

    rr = r0 + len(township_names)
    ws.cell(row=rr, column=1, value="总计/亩")
    colsum = np.sum(matrix_mu, axis=0)
    for j, v in enumerate(colsum, start=2):
        ws.cell(row=rr, column=j, value=int(v))
    ws.cell(row=rr, column=2 + len(class_labels), value=int(np.sum(colsum)))

    ws.column_dimensions["A"].width = 12
    for c in range(2, 3 + len(class_labels)):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 12
    _apply_table_style(ws, 1, 1, rr, 2 + len(class_labels))


def _table3_write_sheet(ws, title, unit, landtypes, class_labels, matrix_mu):
    ws.cell(row=1, column=1, value=title)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2 + len(class_labels))
    ws.cell(row=2, column=1, value="土地利用类型")
    ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
    ws.cell(row=2, column=2, value=f"分级（{unit}）")
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=1 + len(class_labels))
    for j, lab in enumerate(class_labels, start=2):
        ws.cell(row=3, column=j, value=lab)
    ws.cell(row=3, column=2 + len(class_labels), value="总计")

    r0 = 4
    for i, name in enumerate(landtypes):
        rr = r0 + i
        ws.cell(row=rr, column=1, value=name)
        row_vals = matrix_mu[i]
        for j, v in enumerate(row_vals, start=2):
            ws.cell(row=rr, column=j, value=int(v))
        ws.cell(row=rr, column=2 + len(class_labels), value=int(np.sum(row_vals)))

    rr = r0 + len(landtypes)
    ws.cell(row=rr, column=1, value="总计")
    colsum = np.sum(matrix_mu, axis=0)
    for j, v in enumerate(colsum, start=2):
        ws.cell(row=rr, column=j, value=int(v))
    ws.cell(row=rr, column=2 + len(class_labels), value=int(np.sum(colsum)))

    ws.column_dimensions["A"].width = 14
    for c in range(2, 3 + len(class_labels)):
        ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = 12
    _apply_table_style(ws, 1, 1, rr, 2 + len(class_labels))


def _find_latest_out_root(base_out, prefix):
    dirs = [d for d in os.listdir(base_out) if os.path.isdir(os.path.join(base_out, d)) and d.startswith(prefix)]
    if not dirs:
        raise RuntimeError(f"未找到 {prefix} 目录")
    return os.path.join(base_out, sorted(dirs)[-1])

def main():
    arcpy.env.overwriteOutput = True
    arcpy.CheckOutExtension("Spatial")

    base_out = r"c:\Users\HiWin10\Desktop\浮梁自验收\报告及数据修改skill\浮梁属性栅格克里金插值_自验收属性修改20260623"
    out_root = _find_latest_out_root(base_out, prefix="浮梁属性栅格克里金统计_")
    raster_folder = os.path.join(out_root, "02_叠加校正栅格")

    sample_xlsx = os.path.join(base_out, "浮梁土壤属性表层样修正.xlsx")
    standard_xlsx = os.path.join(base_out, "2属性分级标准表.xlsx")
    landuse_shp = os.path.join(base_out, r"fl_landuse\浮梁土地利用.shp")
    township_shp = os.path.join(base_out, r"fl_乡镇界\fl乡镇界.shp")

    out_dir = _ensure_dir(os.path.join(out_root, "06_补充统计表"))
    out1_dir = _ensure_dir(os.path.join(out_dir, "01_土地利用样点-制图对比"))
    out2_dir = _ensure_dir(os.path.join(out_dir, "02_耕地乡镇分级统计"))
    out3_dir = _ensure_dir(os.path.join(out_dir, "03_地类分级统计"))
    gdb = _create_gdb(out_dir, "tables_work.gdb")

    tif_list = [f for f in os.listdir(raster_folder) if f.lower().endswith(".tif")]
    if not tif_list:
        raise RuntimeError("未找到任何校正属性栅格")
    tif_list.sort()
    first_ras = arcpy.Raster(os.path.join(raster_folder, tif_list[0]))
    target_sr = arcpy.Describe(first_ras).spatialReference

    landuse_prj = os.path.join(gdb, "landuse_prj")
    township_prj = os.path.join(gdb, "township_prj")
    if not arcpy.Exists(landuse_prj):
        arcpy.management.Project(landuse_shp, landuse_prj, target_sr)
        _add_landuse_zone_fields(landuse_prj)
    if not arcpy.Exists(township_prj):
        arcpy.management.Project(township_shp, township_prj, target_sr)

    zone_ras, l1_ras = _build_zone_rasters(gdb, landuse_prj, first_ras)
    zone_id_to_label = {
        1: "耕地|水田",
        2: "耕地|水浇地",
        3: "耕地|旱地",
        4: "园地|果园",
        5: "园地|茶园",
        6: "园地|橡胶园",
        7: "园地|其他园地",
        8: "林地",
        9: "草地",
        10: "其他",
    }
    l1_id_to_label = {1: "耕地", 2: "园地", 3: "林地", 4: "草地", 5: "其他"}

    standards, standards_display = _read_standards(standard_xlsx)
    samples = _read_samples(sample_xlsx)
    sidx = samples["header_idx"]

    pub = _published_areas_mu()
    area_rows_order = [
        ("耕地", "水田", "耕地|水田"),
        ("耕地", "水浇地", "耕地|水浇地"),
        ("耕地", "旱地", "耕地|旱地"),
        ("耕地", "合计", "耕地"),
        ("园地", "果园", "园地|果园"),
        ("园地", "茶园", "园地|茶园"),
        ("园地", "橡胶园", "园地|橡胶园"),
        ("园地", "其他园地", "园地|其他园地"),
        ("园地", "合计", "园地"),
        ("林地", "", "林地"),
        ("草地", "", "草地"),
        ("其他", "", "其他"),
        ("全县", "", "行政区总面积"),
    ]

    def _get_unit(attr_key):
        unit_map = {
            "容重": "g/cm³",
            "pH": "",
            "阳离子": "cmol(+)/kg",
            "阳离子交换量": "cmol(+)/kg",
            "有机质": "g/kg",
            "全氮": "g/kg",
            "全磷": "g/kg",
            "全钾": "g/kg",
            "砂粒": "%",
            "粉粒": "%",
            "黏粒": "%",
            "有效磷": "mg/kg",
            "速效钾": "mg/kg",
            "有效铁": "mg/kg",
            "有效锰": "mg/kg",
            "有效铜": "mg/kg",
            "有效锌": "mg/kg",
            "有效硼": "mg/kg",
            "有效钼": "mg/kg",
        }
        return unit_map.get(attr_key, "")

    def _sample_stats_for_attr(attr_key):
        if attr_key not in sidx:
            return {}
        fi = sidx[attr_key]
        li = sidx.get(_norm("土地利用"))
        if li is None:
            return {}
        gvals = {}
        all_vals = []
        for r in samples["rows"]:
            if fi >= len(r) or li >= len(r):
                continue
            v = _safe_float(r[fi])
            if v is None:
                continue
            all_vals.append(v)
            g = _sample_landuse_group(r[li])
            gvals.setdefault(g, []).append(v)
        stats = {}
        for g, vals in gvals.items():
            arr = np.asarray(vals, dtype=float)
            stats[g] = {
                "mean": float(np.mean(arr)) if arr.size else None,
                "min": float(np.min(arr)) if arr.size else None,
                "max": float(np.max(arr)) if arr.size else None,
                "n": int(arr.size),
            }
        if all_vals:
            arr = np.asarray(all_vals, dtype=float)
            stats["全县"] = {
                "mean": float(np.mean(arr)),
                "min": float(np.min(arr)),
                "max": float(np.max(arr)),
                "n": int(arr.size),
            }
        return stats

    def _weighted_mean(means_by_zone, weights_mu, zones):
        num = 0.0
        den = 0.0
        for z in zones:
            m = means_by_zone.get(z)
            w = weights_mu.get(z)
            if m is None or w is None:
                continue
            num += m * float(w)
            den += float(w)
        return None if den <= 0 else num / den

    def _make_table1_for_folder(raster_folder, out_xlsx):
        print(f"开始生成表1: {out_xlsx}")
        wb = openpyxl.Workbook()
        del wb[wb.sheetnames[0]]
        for fn in tif_list:
            attr_key = _norm(fn.replace("_叠加残差克里金.tif", "").replace(".tif", ""))
            rpath = os.path.join(raster_folder, fn)
            if not os.path.exists(rpath):
                continue
            print(f"  表1处理属性: {attr_key}")

            unit = _get_unit(attr_key)
            sstats = _sample_stats_for_attr(attr_key)

            zt = os.path.join(gdb, f"zt_{uuid.uuid4().hex[:8]}")
            means_by_id = _zonal_mean_by_zone_raster(zone_ras, rpath, zt)
            means_by_zone = {zone_id_to_label.get(k, str(k)): v for k, v in means_by_id.items()}
            try:
                if arcpy.Exists(zt):
                    arcpy.management.Delete(zt)
            except Exception:
                pass

            rows = []
            for l1, l2, zkey in area_rows_order:
                if l1 == "全县":
                    s = sstats.get("全县", {})
                    m = _weighted_mean(
                        means_by_zone,
                        {
                            "耕地|水田": pub["耕地|水田"],
                            "耕地|水浇地": pub["耕地|水浇地"],
                            "耕地|旱地": pub["耕地|旱地"],
                            "园地|果园": pub["园地|果园"],
                            "园地|茶园": pub["园地|茶园"],
                            "园地|橡胶园": pub["园地|橡胶园"],
                            "园地|其他园地": pub["园地|其他园地"],
                            "林地": pub["林地"],
                            "草地": pub["草地"],
                            "其他": pub["其他"],
                        },
                        [
                            "耕地|水田",
                            "耕地|水浇地",
                            "耕地|旱地",
                            "园地|果园",
                            "园地|茶园",
                            "园地|橡胶园",
                            "园地|其他园地",
                            "林地",
                            "草地",
                            "其他",
                        ],
                    )
                    area_mu = pub["行政区总面积"]
                    rows.append(
                        {
                            "l1": "全县",
                            "l2": "",
                            "s_mean": _format_mean(s.get("mean")),
                            "s_rng": _format_range(s.get("min"), s.get("max")),
                            "s_n": s.get("n", "—") if s else "—",
                            "m_mean": _format_mean(m),
                            "m_area": _format_area_wanmu(area_mu),
                        }
                    )
                    continue

                if l2 == "合计" and l1 in ("耕地", "园地"):
                    if l1 == "耕地":
                        zones = ["耕地|水田", "耕地|水浇地", "耕地|旱地"]
                    else:
                        zones = ["园地|果园", "园地|茶园", "园地|橡胶园", "园地|其他园地"]
                    sv = []
                    sn = 0
                    smin = None
                    smax = None
                    for z in zones:
                        s0 = sstats.get(z)
                        if not s0:
                            continue
                        sv.append((s0["mean"], s0["n"]))
                        sn += s0["n"]
                        smin = s0["min"] if smin is None else min(smin, s0["min"])
                        smax = s0["max"] if smax is None else max(smax, s0["max"])
                    s_mean = None if sn <= 0 else float(sum(m * n for m, n in sv) / sn)
                    m_mean = _weighted_mean(means_by_zone, pub, zones)
                    area_mu = pub[l1]
                    rows.append(
                        {
                            "l1": l1,
                            "l2": l2,
                            "s_mean": _format_mean(s_mean),
                            "s_rng": _format_range(smin, smax),
                            "s_n": sn if sn > 0 else "—",
                            "m_mean": _format_mean(m_mean),
                            "m_area": _format_area_wanmu(area_mu),
                        }
                    )
                    continue

                if l1 == "其他" and l2 == "":
                    s0 = sstats.get("其他", {})
                    rows.append(
                        {
                            "l1": "其他",
                            "l2": "",
                            "s_mean": _format_mean(s0.get("mean")),
                            "s_rng": _format_range(s0.get("min"), s0.get("max")),
                            "s_n": s0.get("n", "—") if s0 else "—",
                            "m_mean": _format_mean(means_by_zone.get("其他")),
                            "m_area": _format_area_wanmu(pub["其他"]),
                        }
                    )
                    continue

                if l1 in ("林地", "草地") and l2 == "":
                    s0 = sstats.get(l1, {})
                    rows.append(
                        {
                            "l1": l1,
                            "l2": "",
                            "s_mean": _format_mean(s0.get("mean")),
                            "s_rng": _format_range(s0.get("min"), s0.get("max")),
                            "s_n": s0.get("n", "—") if s0 else "—",
                            "m_mean": _format_mean(means_by_zone.get(l1)),
                            "m_area": _format_area_wanmu(pub[l1]),
                        }
                    )
                    continue

                if "|" in zkey:
                    s0 = sstats.get(zkey, {})
                    rows.append(
                        {
                            "l1": l1,
                            "l2": l2,
                            "s_mean": _format_mean(s0.get("mean")),
                            "s_rng": _format_range(s0.get("min"), s0.get("max")),
                            "s_n": s0.get("n", "—") if s0 else "—",
                            "m_mean": _format_mean(means_by_zone.get(zkey)),
                            "m_area": _format_area_wanmu(pub.get(zkey)),
                        }
                    )
                    continue

            sheet_name = attr_key[:31] if attr_key else "sheet"
            if sheet_name in wb.sheetnames:
                sheet_name = (sheet_name[:28] + "_" + str(len(wb.sheetnames)))[:31]
            ws = wb.create_sheet(sheet_name)
            _write_table1_sheet(ws, attr_key, unit, rows)

        wb.save(out_xlsx)

    out_table1 = os.path.join(out1_dir, "表1_土地利用样点与制图对比.xlsx")
    _make_table1_for_folder(raster_folder, out_table1)
    print("表1完成")

    landtypes = ["耕地", "园地", "林地", "草地", "其他"]
    pub_l1 = {k: pub[k] for k in landtypes}
    cropland_target = pub["耕地"]

    wb2 = openpyxl.Workbook()
    del wb2[wb2.sheetnames[0]]
    wb3 = openpyxl.Workbook()
    del wb3[wb3.sheetnames[0]]

    township_field = "TXZQMC" if "TXZQMC" in [f.name for f in arcpy.ListFields(township_prj)] else "XZQMC"

    township_names = []
    with arcpy.da.SearchCursor(township_prj, [township_field]) as cur:
        for (n,) in cur:
            if n is None:
                continue
            township_names.append(str(n))
    township_names = sorted(set(township_names))

    landuse_lyr = "landuse_lyr"
    arcpy.management.MakeFeatureLayer(landuse_prj, landuse_lyr)
    arcpy.management.SelectLayerByAttribute(landuse_lyr, "NEW_SELECTION", "L1GROUP = '耕地'")

    for fn in tif_list:
        attr_key = _norm(fn.replace("_叠加残差克里金.tif", "").replace(".tif", ""))
        rpath = os.path.join(raster_folder, fn)
        if not os.path.exists(rpath):
            continue
        specs = standards.get(attr_key)
        disp = standards_display.get(attr_key)
        if not specs or not disp:
            continue
        print(f"开始处理分级面积: {attr_key}")

        class_labels = [disp.get(i, str(i)) for i in range(1, 6)]
        class_ras = os.path.join(gdb, f"cls_{uuid.uuid4().hex[:8]}")
        _make_class_raster(rpath, specs, class_ras)

        masked_ras = os.path.join(gdb, f"msk_{uuid.uuid4().hex[:8]}")
        Int(SetNull(arcpy.Raster(l1_ras) != 1, arcpy.Raster(class_ras))).save(masked_ras)

        out_ta = os.path.join(gdb, f"ta_{uuid.uuid4().hex[:8]}")
        TabulateArea(township_prj, township_field, masked_ras, "Value", out_ta)

        fields = [f.name for f in arcpy.ListFields(out_ta)]
        value_fields = [f for f in fields if f.lower().startswith("value_")]
        value_fields.sort(key=lambda x: int(x.split("_", 1)[1]))

        raw = {}
        # ArcPy TabulateArea might still create output field named XZQMC or TXZQMC
        out_tf = township_field if township_field in fields else "XZQMC"
        if out_tf not in fields: out_tf = fields[1]  # fallback
        
        with arcpy.da.SearchCursor(out_ta, [out_tf] + value_fields) as cur:
            for row in cur:
                name = row[0]
                if name is None:
                    continue
                areas = [0.0 if v is None else float(v) for v in row[1:]]
                raw[str(name)] = areas

        cell_area_m2 = float(abs(arcpy.Describe(arcpy.Raster(rpath)).meanCellWidth) * abs(arcpy.Describe(arcpy.Raster(rpath)).meanCellHeight))
        to_mu = 1.0 / 666.6666666666666
        mat_mu = []
        for name in township_names:
            a = raw.get(name, [0.0] * len(value_fields))
            mu = [v * to_mu for v in a]
            mat_mu.append(mu[:5] + [0.0] * max(0, 5 - len(mu)))
        mat_mu = np.asarray(mat_mu, dtype=float)[:, :5]

        total_mu = float(np.sum(mat_mu))
        if total_mu > 0:
            factor = float(cropland_target) / total_mu
            mat_scaled = mat_mu * factor
        else:
            mat_scaled = mat_mu

        flat = mat_scaled.flatten()
        rounded = _sum_preserving_round(flat, cropland_target)
        mat_int = rounded.reshape(mat_mu.shape)

        sheet_name = (attr_key + "分级")[:31]
        ws = wb2.create_sheet(sheet_name)
        _table2_write_sheet(ws, f"各乡镇耕地土壤{attr_key}分级面积统计表", "亩", township_names, class_labels, mat_int)

        out_ta2 = os.path.join(gdb, f"ta3_{uuid.uuid4().hex[:8]}")
        TabulateArea(l1_ras, "Value", class_ras, "Value", out_ta2)
        fields = [f.name for f in arcpy.ListFields(out_ta2)]
        value_fields = [f for f in fields if f.lower().startswith("value_")]
        value_fields.sort(key=lambda x: int(x.split("_", 1)[1]))

        l1_raw = {}
        with arcpy.da.SearchCursor(out_ta2, ["Value"] + value_fields) as cur:
            for row in cur:
                z = row[0]
                if z is None:
                    continue
                areas = [0.0 if v is None else float(v) for v in row[1:]]
                l1_raw[l1_id_to_label.get(int(z), str(z))] = areas

        mat_l1_mu = []
        for lt in landtypes:
            a = l1_raw.get(lt, [0.0] * len(value_fields))
            mu = [v * to_mu for v in a]
            row = (mu[:5] + [0.0] * max(0, 5 - len(mu)))[:5]
            mat_l1_mu.append(row)
        mat_l1_mu = np.asarray(mat_l1_mu, dtype=float)

        mat_l1_int = np.zeros_like(mat_l1_mu, dtype=int)
        for i, lt in enumerate(landtypes):
            target = int(pub_l1[lt])
            row = mat_l1_mu[i]
            s = float(np.sum(row))
            if s <= 0:
                mat_l1_int[i] = 0
                continue
            scaled = row * (target / s)
            mat_l1_int[i] = _sum_preserving_round(scaled, target)

        ws = wb3.create_sheet(sheet_name)
        _table3_write_sheet(ws, f"各地类土壤{attr_key}分级面积统计表", "亩", landtypes, class_labels, mat_l1_int)

        for p in [class_ras, masked_ras, out_ta, out_ta2]:
            try:
                if arcpy.Exists(p):
                    arcpy.management.Delete(p)
            except Exception:
                pass

    out2_xlsx = os.path.join(out2_dir, "表2_耕地_乡镇分级统计.xlsx")
    wb2.save(out2_xlsx)
    out3_xlsx = os.path.join(out3_dir, "表3_各地类分级统计.xlsx")
    wb3.save(out3_xlsx)

    meta_xlsx = os.path.join(out_dir, "meta_补充统计表.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "meta"
    meta = {
        "out_root": out_root,
        "raster_folder": raster_folder,
        "sample_xlsx": sample_xlsx,
        "standard_xlsx": standard_xlsx,
        "landuse_shp": landuse_shp,
        "township_shp": township_shp,
        "time": dt.datetime.now().isoformat(timespec="seconds"),
        "published_total_mu": pub["行政区总面积"],
        "published_cropland_mu": pub["耕地"],
        "published_garden_mu": pub["园地"],
        "published_forest_mu": pub["林地"],
        "published_grass_mu": pub["草地"],
        "published_other_mu": pub["其他"],
    }
    rr = 1
    for k, v in meta.items():
        ws.cell(row=rr, column=1, value=k)
        ws.cell(row=rr, column=2, value=v)
        rr += 1
    wb.save(meta_xlsx)

    print("补充统计表完成")
    print(out_dir)


if __name__ == "__main__":
    main()
