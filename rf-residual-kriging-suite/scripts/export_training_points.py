import os
import re
import math
import uuid
import arcpy
from arcpy.sa import ExtractValuesToPoints
import openpyxl

def _norm(s):
    if s is None:
        return ""
    s = str(s)
    s = s.replace("\u3000", " ")
    s = s.replace(" ", "")
    s = s.strip()
    return s

def _safe_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        try:
            x = float(v)
            if math.isnan(x) or math.isinf(x):
                return None
            return x
        except Exception:
            return None
    try:
        s = str(v).strip()
        if s == "":
            return None
        x = float(s)
        if math.isnan(x) or math.isinf(x):
            return None
        return x
    except Exception:
        return None

def _ensure_dir(p):
    os.makedirs(p, exist_ok=True)
    return p

def _create_gdb(folder, name):
    gdb = os.path.join(folder, name)
    if not arcpy.Exists(gdb):
        arcpy.management.CreateFileGDB(folder, name)
    return gdb

def _find_latest_out_root(base_folder, prefix="RF_残差克里金校正成果_"):
    dirs = [d for d in os.listdir(base_folder) if os.path.isdir(os.path.join(base_folder, d)) and d.startswith(prefix)]
    if not dirs:
        raise RuntimeError(f"未在目录中找到 {prefix} 文件夹")
    dirs.sort()
    return os.path.join(base_folder, dirs[-1])

def _read_validation_lonlat(validation_xlsx):
    wb = openpyxl.load_workbook(validation_xlsx, data_only=True, read_only=True)
    if "验证集完整数据" in wb.sheetnames:
        ws = wb["验证集完整数据"]
    elif "合并验证集" in wb.sheetnames:
        ws = wb["合并验证集"]
    else:
        ws = wb[wb.sheetnames[0]]
    head = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {_norm(h): i for i, h in enumerate(head) if h is not None}
    need = ["属性", "经度", "纬度"]
    for k in need:
        if k not in idx:
            raise RuntimeError(f"验证集表缺少字段: {k}")

    out = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r:
            continue
        attr = _norm(r[idx["属性"]])
        if attr == "":
            continue
        lon = _safe_float(r[idx["经度"]])
        lat = _safe_float(r[idx["纬度"]])
        if lon is None or lat is None:
            continue
        # Use round(val, 5) to match coordinates safely
        out.setdefault(attr, set()).add((round(lon, 5), round(lat, 5)))
    return out

def _choose_sheet(wb):
    # 优先选择包含特定关键词的表
    preferred_keywords = ["数据清洗", "修正后", "clean", "final"]
    for name in wb.sheetnames:
        if any(keyword in name for keyword in preferred_keywords):
            return name
    
    # 如果没有找到优先表，再选择行数最多的
    best = None
    best_count = -1
    for name in wb.sheetnames:
        ws = wb[name]
        count = 0
        for r in ws.iter_rows(min_row=2, values_only=True):
            if r and any(v is not None and str(v).strip() != "" for v in r):
                count += 1
            if count >= 5000:
                break
        if count > best_count:
            best = name
            best_count = count
    return best

def _read_full_samples(sample_xlsx, valid_attrs):
    wb = openpyxl.load_workbook(sample_xlsx, data_only=True, read_only=True)
    sheet = _choose_sheet(wb)
    ws = wb[sheet]
    head = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {_norm(h): i for i, h in enumerate(head) if h is not None and _norm(h) != ""}
    
    lon_keys = ["longitude", "lon", "x", "jd", "经度", "定位经度", "布设经度"]
    lat_keys = ["latitude", "lat", "y", "wd", "纬度", "定位纬度", "定位维度", "布设纬度"]
    lon_i = None
    lat_i = None
    for k in lon_keys:
        if k in idx:
            lon_i = idx[k]
            break
    for k in lat_keys:
        if k in idx:
            lat_i = idx[k]
            break
    
    if lon_i is None or lat_i is None:
        raise RuntimeError(f"样点表缺少经纬度字段。当前表头: {list(idx.keys())}")

    out = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r:
            continue
        lon = _safe_float(r[lon_i] if lon_i < len(r) else None)
        lat = _safe_float(r[lat_i] if lat_i < len(r) else None)
        if lon is None or lat is None:
            continue
            
        for attr in valid_attrs:
            attr_i = idx.get(attr)
            if attr_i is not None and attr_i < len(r):
                actual = _safe_float(r[attr_i])
                if actual is not None:
                    out.setdefault(attr, []).append((lon, lat, actual))
    return out

def _list_corrected_rasters(folder):
    out = {}
    for fn in os.listdir(folder):
        if not fn.lower().endswith(".tif"):
            continue
        if "_叠加残差克里金" not in fn:
            continue
        base = fn.replace("_叠加残差克里金.tif", "").replace("_叠加残差克里金.TIF", "")
        out[_norm(base)] = os.path.join(folder, fn)
    return out

def _export_one_attr(gdb, out_folder, attr, lonlat_actual_rows, raster_path):
    akey = _norm(attr)
    r = arcpy.Raster(raster_path)
    sr = arcpy.Describe(r).spatialReference
    gcs = arcpy.SpatialReference(4490)

    pts_gcs = os.path.join(gdb, f"t_{uuid.uuid4().hex[:8]}")
    arcpy.management.CreateFeatureclass(gdb, os.path.basename(pts_gcs), "POINT", spatial_reference=gcs)
    arcpy.management.AddField(pts_gcs, "lon", "DOUBLE")
    arcpy.management.AddField(pts_gcs, "lat", "DOUBLE")
    arcpy.management.AddField(pts_gcs, "actual", "DOUBLE")

    with arcpy.da.InsertCursor(pts_gcs, ["SHAPE@XY", "lon", "lat", "actual"]) as ic:
        for lon, lat, actual in lonlat_actual_rows:
            ic.insertRow([(lon, lat), lon, lat, actual])

    pts = pts_gcs
    if sr and sr.type != "Geographic":
        pts_prj = os.path.join(gdb, f"tp_{uuid.uuid4().hex[:8]}")
        arcpy.management.Project(pts_gcs, pts_prj, sr)
        pts = pts_prj

    arcpy.env.snapRaster = r
    arcpy.env.extent = r
    arcpy.env.cellSize = r

    out_pts = os.path.join(gdb, f"ex_{uuid.uuid4().hex[:8]}")
    ExtractValuesToPoints(pts, r, out_pts, interpolate_values="INTERPOLATE", add_attributes="VALUE_ONLY")

    rows = []
    with arcpy.da.SearchCursor(out_pts, ["lon", "lat", "actual", "RASTERVALU"]) as cur:
        for lon, lat, actual, pred in cur:
            if actual is None or lon is None or lat is None:
                continue
            pred2 = _safe_float(pred)
            resid = None if pred2 is None else float(actual) - pred2
            rows.append((float(lon), float(lat), float(actual), pred2, resid))

    out_xlsx = os.path.join(out_folder, f"{akey}_训练集点_叠加残差克里金.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "training_points"
    ws.append(["经度", "纬度", "实际值", "预测值(叠加残差克里金)", "残差(实际-预测)"])
    for lon, lat, actual, pred2, resid in rows:
        ws.append([lon, lat, actual, pred2, resid])
    wb.save(out_xlsx)

    for p in [out_pts, pts_gcs]:
        try:
            if arcpy.Exists(p):
                arcpy.management.Delete(p)
        except Exception:
            pass
    if pts != pts_gcs:
        try:
            if arcpy.Exists(pts):
                arcpy.management.Delete(pts)
        except Exception:
            pass

    return out_xlsx, len(rows)

def main():
    arcpy.env.overwriteOutput = True
    arcpy.CheckOutExtension("Spatial")

    work_dir = r"c:\Users\HiWin10\Desktop\浮梁自验收\报告及数据修改skill\浮梁属性栅格克里金插值_自验收属性修改20260623"
    rf_folder = os.path.join(work_dir, r"预测属性栅格")
    validation_xlsx = os.path.join(work_dir, "浮梁_验证集-随机森林最优方案样本验证集.xlsx")
    sample_xlsx = os.path.join(work_dir, "浮梁土壤属性表层样修正.xlsx")

    out_root = _find_latest_out_root(work_dir, prefix="浮梁属性栅格克里金统计_")
    
    # 尝试读取下限校正的栅格目录，如果不存在则使用普通叠加校正目录
    corr_folder_lb = os.path.join(out_root, "02_叠加校正栅格_下限校正")
    corr_folder_pre = os.path.join(out_root, "02_叠加校正栅格")
    corr_folder = corr_folder_lb if os.path.exists(corr_folder_lb) else corr_folder_pre
    
    out_folder = _ensure_dir(os.path.join(out_root, "05_训练集点表_下限校正"))
    gdb = _create_gdb(out_root, "export_train_work.gdb")

    print("1. 读取验证集点坐标...")
    val_lonlat = _read_validation_lonlat(validation_xlsx)
    
    rasters = _list_corrected_rasters(corr_folder)
    valid_attrs = list(rasters.keys())

    print("2. 从总样点表中读取训练集...")
    full_samples = _read_full_samples(sample_xlsx, valid_attrs)
    
    train_samples = {}
    for attr, pts in full_samples.items():
        v_set = val_lonlat.get(attr, set())
        t_list = []
        for lon, lat, actual in pts:
            # 过滤掉属于验证集的点
            if (round(lon, 5), round(lat, 5)) not in v_set:
                t_list.append((lon, lat, actual))
        train_samples[attr] = t_list

    done = []
    for attr, pts in train_samples.items():
        rp = rasters.get(attr)
        if rp is None:
            continue
        print(f"正在导出 {attr} 的训练集点...")
        out_xlsx, n = _export_one_attr(gdb, out_folder, attr, pts, rp)
        done.append((attr, out_xlsx, n))

    summary_xlsx = os.path.join(out_folder, "训练集点_叠加残差克里金_导出清单.xlsx")
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "list"
    ws.append(["属性", "训练集记录数", "文件"])
    for a, p, n in sorted(done, key=lambda x: x[0]):
        ws.append([a, n, p])
    wb.save(summary_xlsx)

    print("训练集点导出完成。结果保存在:")
    print(out_folder)

if __name__ == "__main__":
    main()