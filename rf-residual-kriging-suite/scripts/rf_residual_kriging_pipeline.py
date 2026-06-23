import os
import re
import math
import uuid
import datetime as dt

import arcpy
from arcpy.sa import (
    Kriging,
    KrigingModelOrdinary,
    RadiusVariable,
    ExtractValuesToPoints,
    Plus,
    Con,
    SetNull,
    Int,
)

import numpy as np
import openpyxl


def _lower_bounds_from_png_table():
    return {
        _norm("容重"): 0.5,
        _norm("pH"): 3.5,
        _norm("阳离子"): 0.5,
        _norm("阳离子交换量"): 0.5,
        _norm("有机质"): 1.0,
        _norm("全氮"): 0.1,
        _norm("全磷"): 0.1,
        _norm("全钾"): 3.0,
        _norm("有效磷"): 0.25,
        _norm("速效钾"): 10.0,
        _norm("有效铁"): 1.0,
        _norm("有效锰"): 1.0,
        _norm("有效铜"): 0.1,
        _norm("有效锌"): 0.1,
        _norm("有效硼"): 0.1,
        _norm("有效钼"): 0.01,
    }


def _clamp_raster_below_zero_inplace(raster_path, attr_key, lower_bounds):
    lb = lower_bounds.get(attr_key)
    if lb is None:
        return None
    r = arcpy.Raster(raster_path)
    try:
        arcpy.management.CalculateStatistics(r)
    except Exception:
        pass
    try:
        vmin = float(arcpy.management.GetRasterProperties(r, "MINIMUM").getOutput(0))
    except Exception:
        return None
    if vmin < 0:
        Con(r < 0, lb, r).save(raster_path)
        return {"min": vmin, "lower_bound": lb}
    return None


def _norm(s):
    if s is None:
        return ""
    s = str(s)
    s = s.replace("\u3000", " ")
    s = s.replace(" ", "")
    s = s.strip()
    return s


def _safe_float(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        try:
            x = float(v)
            if math.isnan(x) or math.isinf(x):
                return None
            return x
        except Exception:
            return None
    try:
        s = str(v).strip()
        if s == "":
            return None
        x = float(s)
        if math.isnan(x) or math.isinf(x):
            return None
        return x
    except Exception:
        return None


def _r2_rmse(y_true, y_pred):
    y_true = np.asarray(y_true, dtype=float)
    y_pred = np.asarray(y_pred, dtype=float)
    if y_true.size == 0:
        return None, None
    mse = float(np.mean((y_true - y_pred) ** 2))
    rmse = float(math.sqrt(mse))
    mean_y = float(np.mean(y_true))
    ss_tot = float(np.sum((y_true - mean_y) ** 2))
    ss_res = float(np.sum((y_true - y_pred) ** 2))
    r2 = None if ss_tot <= 0 else 1.0 - (ss_res / ss_tot)
    return r2, rmse


def _choose_sheet(wb):
    # 优先选择包含特定关键词的表
    preferred_keywords = ["数据清洗", "修正后", "clean", "final"]
    for name in wb.sheetnames:
        if any(keyword in name for keyword in preferred_keywords):
            return name
    
    # 如果没有找到优先表，再选择行数最多的
    best = None
    best_count = -1
    for name in wb.sheetnames:
        ws = wb[name]
        count = 0
        for r in ws.iter_rows(min_row=2, values_only=True):
            if r and any(v is not None and str(v).strip() != "" for v in r):
                count += 1
            if count >= 5000:
                break
        if count > best_count:
            best = name
            best_count = count
    return best


def _parse_range_spec(spec):
    if spec is None:
        return []
    s = str(spec).strip()
    if s == "":
        return []
    s = s.replace("≤", "<=").replace("≥", ">=")
    s = s.replace("，", ",")
    tokens = []
    tokens.extend(re.findall(r"[\(\[]\s*[-\d.]+\s*,\s*[-\d.]+\s*[\)\]]", s))
    tokens.extend(re.findall(r"(?:<=|>=|<|>)\s*[-\d.]+", s))
    out = []
    for t in tokens:
        t = t.replace(" ", "")
        if t.startswith(("(", "[")) and t.endswith((")", "]")):
            left_inc = t[0] == "["
            right_inc = t[-1] == "]"
            body = t[1:-1]
            a, b = body.split(",", 1)
            lo = _safe_float(a)
            hi = _safe_float(b)
            if lo is None or hi is None:
                continue
            out.append((lo, hi, left_inc, right_inc))
        else:
            m = re.match(r"(<=|>=|<|>)\s*([-\d.]+)", t)
            if not m:
                continue
            op = m.group(1)
            x = _safe_float(m.group(2))
            if x is None:
                continue
            if op == ">":
                out.append((x, None, False, False))
            elif op == ">=":
                out.append((x, None, True, False))
            elif op == "<":
                out.append((None, x, False, False))
            elif op == "<=":
                out.append((None, x, False, True))
    return out


def _classify_value(v, class_specs):
    x = _safe_float(v)
    if x is None:
        return None
    for cls in sorted(class_specs.keys()):
        segs = class_specs[cls]
        for lo, hi, lo_inc, hi_inc in segs:
            ok = True
            if lo is not None:
                ok = ok and (x >= lo if lo_inc else x > lo)
            if hi is not None:
                ok = ok and (x <= hi if hi_inc else x < hi)
            if ok:
                return cls
    return None


def _segs_to_raster_cond(r, segs):
    cond = None
    for lo, hi, lo_inc, hi_inc in segs:
        c = None
        if lo is not None:
            c = (r >= lo) if lo_inc else (r > lo)
        if hi is not None:
            c2 = (r <= hi) if hi_inc else (r < hi)
            c = c2 if c is None else (c & c2)
        if c is None:
            continue
        cond = c if cond is None else (cond | c)
    return cond


def _read_standards(standard_xlsx):
    wb = openpyxl.load_workbook(standard_xlsx, data_only=False)
    ws = wb[wb.sheetnames[0]]
    headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    attr_cols = {}
    for c in range(3, ws.max_column + 1):
        v = headers[c - 1]
        if v is None:
            continue
        attr = _norm(v)
        if attr != "":
            attr_cols[attr] = c

    standards = {}
    display = {}
    for cls_row in range(2, 7):
        cls = ws.cell(row=cls_row, column=2).value
        cls = int(cls) if cls is not None else None
        if cls is None:
            continue
        for attr, col in attr_cols.items():
            v = ws.cell(row=cls_row, column=col).value
            if v is None or str(v).strip() == "":
                continue
            standards.setdefault(attr, {}).setdefault(cls, []).extend(_parse_range_spec(v))
            display.setdefault(attr, {})[cls] = str(v).strip()

    return standards, display


def _read_validation_long(validation_xlsx):
    wb = openpyxl.load_workbook(validation_xlsx, data_only=True, read_only=True)
    out = {}
    
    if "验证集完整数据" in wb.sheetnames:
        sheets_to_read = [wb["验证集完整数据"]]
    elif "合并验证集" in wb.sheetnames:
        sheets_to_read = [wb["合并验证集"]]
    else:
        sheets_to_read = wb.worksheets
    for ws in sheets_to_read:
        head_row = next(ws.iter_rows(min_row=1, max_row=1), None)
        if not head_row: continue
        head = [c.value for c in head_row]
        idx = {_norm(h): i for i, h in enumerate(head) if h is not None}
        need = ["属性", "实际值", "预测值", "残差", "经度", "纬度"]
        if any(k not in idx for k in need):
            continue
            
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r:
                continue
            attr = _norm(r[idx["属性"]])
            if attr == "":
                continue
            actual = _safe_float(r[idx["实际值"]])
            pred = _safe_float(r[idx["预测值"]])
            resid = _safe_float(r[idx["残差"]])
            lon = _safe_float(r[idx["经度"]])
            lat = _safe_float(r[idx["纬度"]])
            if lon is None or lat is None:
                continue
            if actual is None or pred is None:
                continue
            if resid is None:
                resid = actual - pred
            out.setdefault(attr, []).append((lon, lat, actual, pred, resid))
    return out


def _read_samples(sample_xlsx):
    wb = openpyxl.load_workbook(sample_xlsx, data_only=True, read_only=True)
    sheet = _choose_sheet(wb)
    ws = wb[sheet]
    head = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {_norm(h): i for i, h in enumerate(head) if h is not None and _norm(h) != ""}
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r:
            continue
        lon = _safe_float(r[idx.get("longitude", -1)]) if "longitude" in idx else None
        lat = _safe_float(r[idx.get("latitude", -1)]) if "latitude" in idx else None
        if lon is None or lat is None:
            continue
        rows.append(r)
    return {"sheet": sheet, "header_idx": idx, "rows": rows}


def _list_rf_rasters(folder):
    out = {}
    for fn in os.listdir(folder):
        if not fn.lower().endswith(".tif"):
            continue
        base = fn.replace("_预测结果.tif", "").replace("_预测结果.TIF", "").replace(".tif", "").replace(".TIF", "")
        out[_norm(base)] = os.path.join(folder, fn)
    return out


def _ensure_dir(p):
    os.makedirs(p, exist_ok=True)
    return p


def _create_gdb(folder, name):
    gdb = os.path.join(folder, name)
    if not arcpy.Exists(gdb):
        arcpy.management.CreateFileGDB(folder, name)
    return gdb


def _make_points_fc(gdb, name, spatial_ref, fields):
    fc = os.path.join(gdb, name)
    arcpy.management.CreateFeatureclass(gdb, name, "POINT", spatial_reference=spatial_ref)
    for fname, ftype in fields:
        arcpy.management.AddField(fc, fname, ftype)
    return fc


def _write_accuracy_xlsx(path, accuracy_rows, steps_text, meta):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "精度对比"
    ws.append(["属性", "验证点数", "R2_RF", "RMSE_RF", "R2_叠加残差克里金", "RMSE_叠加残差克里金"])
    for r in accuracy_rows:
        ws.append([r["attr"], r["n"], r["r2_rf"], r["rmse_rf"], r["r2_corr"], r["rmse_corr"]])

    ws2 = wb.create_sheet("处理步骤")
    for i, line in enumerate(steps_text.splitlines(), start=1):
        ws2.cell(row=i, column=1, value=line)

    ws3 = wb.create_sheet("meta")
    rr = 1
    for k, v in meta.items():
        ws3.cell(row=rr, column=1, value=k)
        ws3.cell(row=rr, column=2, value=v)
        rr += 1

    wb.save(path)


def _apply_table_style(ws, top, left, bottom, right):
    from openpyxl.styles import Border, Side, Alignment, Font

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(top, bottom + 1):
        for c in range(left, right + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.cell(row=top, column=left).font = Font(bold=True)


def _write_grade_sheet(wb, attr, class_display, sample_stats, map_stats, sample_mm, map_mm):
    name = f"{attr}分级"
    if len(name) > 31:
        name = name[:31]
    ws = wb.create_sheet(name)

    ws.cell(row=1, column=1, value=f"土壤{attr}分级分布")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)

    ws.cell(row=2, column=1, value="土壤三普分级")
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
    ws.cell(row=2, column=3, value="样点统计")
    ws.merge_cells(start_row=2, start_column=3, end_row=2, end_column=4)
    ws.cell(row=2, column=5, value="制图统计")
    ws.merge_cells(start_row=2, start_column=5, end_row=2, end_column=6)

    ws.cell(row=3, column=1, value="分级")
    ws.cell(row=3, column=2, value="值域")
    ws.cell(row=3, column=3, value="数量/个")
    ws.cell(row=3, column=4, value="占比/%")
    ws.cell(row=3, column=5, value="面积/万亩")
    ws.cell(row=3, column=6, value="占比/%")

    row0 = 4
    total_n = sample_stats["total_n"]
    total_area = map_stats["total_area_wanmu"]
    for cls in range(1, 6):
        ws.cell(row=row0 + cls - 1, column=1, value=cls)
        ws.cell(row=row0 + cls - 1, column=2, value=class_display.get(cls, ""))
        n = sample_stats["counts"].get(cls, 0)
        p = 0.0 if total_n <= 0 else (n * 100.0 / total_n)
        ws.cell(row=row0 + cls - 1, column=3, value=n)
        ws.cell(row=row0 + cls - 1, column=4, value=round(p, 2))
        a = map_stats["areas_wanmu"].get(cls, 0.0)
        ap = 0.0 if total_area <= 0 else (a * 100.0 / total_area)
        ws.cell(row=row0 + cls - 1, column=5, value=round(a, 2))
        ws.cell(row=row0 + cls - 1, column=6, value=round(ap, 2))

    r_total = row0 + 5
    ws.cell(row=r_total, column=1, value="全县")
    ws.merge_cells(start_row=r_total, start_column=1, end_row=r_total, end_column=2)
    ws.cell(row=r_total, column=3, value=total_n)
    ws.cell(row=r_total, column=4, value=100 if total_n > 0 else 0)
    ws.cell(row=r_total, column=5, value=round(total_area, 2))
    ws.cell(row=r_total, column=6, value=100 if total_area > 0 else 0)

    r_mean = r_total + 1
    ws.cell(row=r_mean, column=1, value="全县均值")
    ws.merge_cells(start_row=r_mean, start_column=1, end_row=r_mean, end_column=2)
    ws.cell(row=r_mean, column=3, value=sample_mm["mean"])
    ws.merge_cells(start_row=r_mean, start_column=3, end_row=r_mean, end_column=4)
    ws.cell(row=r_mean, column=5, value=map_mm["mean"])
    ws.merge_cells(start_row=r_mean, start_column=5, end_row=r_mean, end_column=6)

    r_median = r_total + 2
    ws.cell(row=r_median, column=1, value="全县中位值")
    ws.merge_cells(start_row=r_median, start_column=1, end_row=r_median, end_column=2)
    ws.cell(row=r_median, column=3, value=sample_mm["median"])
    ws.merge_cells(start_row=r_median, start_column=3, end_row=r_median, end_column=4)
    ws.cell(row=r_median, column=5, value=map_mm["median"])
    ws.merge_cells(start_row=r_median, start_column=5, end_row=r_median, end_column=6)

    r_range = r_total + 3
    ws.cell(row=r_range, column=1, value="全县范围")
    ws.merge_cells(start_row=r_range, start_column=1, end_row=r_range, end_column=2)
    ws.cell(row=r_range, column=3, value=sample_mm["range"])
    ws.merge_cells(start_row=r_range, start_column=3, end_row=r_range, end_column=4)
    ws.cell(row=r_range, column=5, value=map_mm["range"])
    ws.merge_cells(start_row=r_range, start_column=5, end_row=r_range, end_column=6)

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[3].height = 30
    ws.column_dimensions["A"].width = 10
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 10
    _apply_table_style(ws, 1, 1, r_range, 6)


def _calc_map_class_area_stats(class_raster_path):
    arcpy.management.BuildRasterAttributeTable(class_raster_path, "Overwrite")
    r = arcpy.Raster(class_raster_path)
    desc = arcpy.Describe(class_raster_path)
    cell_area = float(abs(desc.meanCellWidth) * abs(desc.meanCellHeight))
    counts = {}
    fields = arcpy.ListFields(class_raster_path)
    f_value = None
    f_count = None
    for f in fields:
        n = f.name.lower()
        if n == "value":
            f_value = f.name
        elif n == "count":
            f_count = f.name
    if not f_value or not f_count:
        raise RuntimeError("分级栅格属性表缺少 Value/Count 字段")

    with arcpy.da.SearchCursor(class_raster_path, [f_value, f_count]) as cur:
        for v, c in cur:
            if v is None or c is None:
                continue
            iv = int(v)
            if iv < 1 or iv > 5:
                continue
            counts[iv] = int(c)
    areas_wanmu = {}
    total = 0.0
    for cls, cnt in counts.items():
        a_m2 = cnt * cell_area
        a_wanmu = a_m2 / 6666666.666666667
        areas_wanmu[cls] = float(a_wanmu)
        total += float(a_wanmu)
    return {"counts": counts, "areas_wanmu": areas_wanmu, "total_area_wanmu": float(total)}


def _calc_raster_stats_with_median(raster_path, gdb, seed, n_points=3000):
    r = arcpy.Raster(raster_path)
    try:
        arcpy.management.CalculateStatistics(r)
    except Exception:
        pass

    def _prop(p):
        try:
            return float(arcpy.management.GetRasterProperties(r, p).getOutput(0))
        except Exception:
            return None

    vmin = _prop("MINIMUM")
    vmax = _prop("MAXIMUM")
    vmean = _prop("MEAN")

    desc = arcpy.Describe(r)
    ext = desc.extent
    rp = os.path.join(gdb, f"rp_{uuid.uuid4().hex[:8]}")
    outp = os.path.join(gdb, f"rv_{uuid.uuid4().hex[:8]}")
    try:
        arcpy.management.CreateRandomPoints(
            os.path.dirname(rp),
            os.path.basename(rp),
            constraining_extent=f"{ext.XMin} {ext.YMin} {ext.XMax} {ext.YMax}",
            number_of_points_or_field=n_points,
        )
        ExtractValuesToPoints(rp, r, outp, interpolate_values="NONE", add_attributes="VALUE_ONLY")
        vals = []
        with arcpy.da.SearchCursor(outp, ["RASTERVALU"]) as cur:
            for vv, in cur:
                if vv is None:
                    continue
                try:
                    x = float(vv)
                except Exception:
                    continue
                if math.isnan(x) or math.isinf(x):
                    continue
                vals.append(x)
        median = None if len(vals) == 0 else float(np.median(np.asarray(vals, dtype=float)))
    finally:
        for p in [rp, outp]:
            try:
                if arcpy.Exists(p):
                    arcpy.management.Delete(p)
            except Exception:
                pass

    return {
        "min": vmin,
        "max": vmax,
        "mean": vmean,
        "median": median,
        "range": None if vmin is None or vmax is None else f"{round(vmin, 2)}~{round(vmax, 2)}",
    }


def _calc_sample_stats(values, class_specs):
    vals = np.asarray([v for v in values if v is not None], dtype=float)
    if vals.size == 0:
        mm = {"mean": None, "median": None, "range": None}
    else:
        mm = {
            "mean": float(np.mean(vals)),
            "median": float(np.median(vals)),
            "range": f"{round(float(np.min(vals)), 2)}~{round(float(np.max(vals)), 2)}",
        }

    counts = {i: 0 for i in range(1, 6)}
    n = 0
    for v in values:
        cls = _classify_value(v, class_specs)
        if cls is None:
            continue
        counts[cls] += 1
        n += 1
    return {"counts": counts, "total_n": n, "mm": mm}


def main():
    arcpy.env.overwriteOutput = True
    arcpy.CheckOutExtension("Spatial")

    base_dir = r"c:\Users\HiWin10\Desktop\浮梁自验收\报告及数据修改skill\浮梁属性栅格克里金插值_自验收属性修改20260623"
    rf_folder = os.path.join(base_dir, r"预测属性栅格")
    validation_xlsx = os.path.join(base_dir, r"浮梁_验证集-随机森林最优方案样本验证集.xlsx")
    sample_xlsx = os.path.join(base_dir, r"浮梁土壤属性表层样修正.xlsx")
    standard_xlsx = os.path.join(base_dir, r"2属性分级标准表.xlsx")

    run_tag = dt.datetime.now().strftime("%Y%m%d_%H%M%S")
    out_root = _ensure_dir(os.path.join(base_dir, f"浮梁属性栅格克里金统计_{run_tag}"))
    out_resid = _ensure_dir(os.path.join(out_root, "01_残差克里金栅格"))
    out_corr = _ensure_dir(os.path.join(out_root, "02_叠加校正栅格"))
    out_class = _ensure_dir(os.path.join(out_root, "03_分级栅格"))
    out_tbl = _ensure_dir(os.path.join(out_root, "04_表格"))
    gdb = _create_gdb(out_root, "work.gdb")

    standards, standards_display = _read_standards(standard_xlsx)
    val_by_attr = _read_validation_long(validation_xlsx)
    sample = _read_samples(sample_xlsx)
    rf_rasters = _list_rf_rasters(rf_folder)
    lower_bounds = _lower_bounds_from_png_table()

    steps_text = "\n".join(
        [
            "1) 准备数据：RF预测栅格、验证集点(含实际值/预测值/残差)、分级标准表、样点表。",
            "2) 以验证集残差作为Z值，使用普通克里金(Kriging)对残差点进行插值，输出残差栅格；环境设置与RF栅格一致（snapRaster/extent/cellSize）。",
            "3) 用栅格计算器将 RF预测栅格 + 残差克里金栅格，得到叠加校正后的属性栅格。",
            "4) 精度验证：对验证集点提取叠加校正栅格值，与实际值计算R²与RMSE；RF原始精度用验证集表中的实际值/预测值计算。",
            "5) 分级制图：按分级标准表对叠加校正栅格重分类得到1~5级分级栅格（支持多区间）。",
            "6) 统计对比：样点按同一分级标准计算各级数量/占比；分级栅格统计各级面积/占比（面积单位万亩）。",
            "7) 输出成果：残差栅格、叠加校正栅格、分级栅格、精度对比表、分级对比表。",
        ]
    )

    accuracy_rows = []
    wb_out = openpyxl.Workbook()
    del wb_out[wb_out.sheetnames[0]]

    gcs = arcpy.SpatialReference(4490)
    semivariogram = "SPHERICAL"
    model = KrigingModelOrdinary(semivariogram)
    search_radius = RadiusVariable(12)
    seed = 20260406

    for attr, rows in val_by_attr.items():
        akey = _norm(attr)
        rf_path = rf_rasters.get(akey)
        if rf_path is None:
            continue

        r_rf = arcpy.Raster(rf_path)
        sr = arcpy.Describe(r_rf).spatialReference
        cell_size = float(arcpy.management.GetRasterProperties(r_rf, "CELLSIZEX").getOutput(0))
        if not (cell_size > 0):
            cell_size = None

        pts_gcs = _make_points_fc(gdb, f"pts_{uuid.uuid4().hex[:8]}", gcs, [("actual", "DOUBLE"), ("pred", "DOUBLE"), ("resid", "DOUBLE")])
        with arcpy.da.InsertCursor(pts_gcs, ["SHAPE@XY", "actual", "pred", "resid"]) as ic:
            for lon, lat, actual, pred, resid in rows:
                ic.insertRow([(lon, lat), actual, pred, resid])

        pts = pts_gcs
        if sr and sr.type != "Geographic":
            pts_prj = os.path.join(gdb, f"ptsprj_{uuid.uuid4().hex[:8]}")
            arcpy.management.Project(pts_gcs, pts_prj, sr)
            pts = pts_prj

        arcpy.env.snapRaster = r_rf
        arcpy.env.extent = r_rf
        arcpy.env.cellSize = r_rf

        resid_ras_path = os.path.join(out_resid, f"{akey}_残差克里金.tif")
        corr_ras_path = os.path.join(out_corr, f"{akey}_叠加残差克里金.tif")
        class_ras_path = os.path.join(out_class, f"{akey}_叠加残差克里金_分级.tif")

        resid_obj = Kriging(pts, "resid", model, cell_size, search_radius)
        resid_obj.save(resid_ras_path)

        from arcpy.sa import Plus, Con
        corr_obj = Plus(r_rf, arcpy.Raster(resid_ras_path))
        
        lb = lower_bounds.get(akey)
        clamp_info = None
        if lb is not None:
            # To check if there are <0 values without saving, we can just save it first to a temp, or save directly and CalculateStatistics
            pass
            
        corr_ras_temp = os.path.join(out_corr, f"tmp_{uuid.uuid4().hex[:8]}.tif")
        corr_obj.save(corr_ras_temp)
        
        r_temp = arcpy.Raster(corr_ras_temp)
        try:
            arcpy.management.CalculateStatistics(r_temp)
            vmin = float(arcpy.management.GetRasterProperties(r_temp, "MINIMUM").getOutput(0))
            if lb is not None and vmin < 0:
                clamp_info = {"min": vmin, "lower_bound": lb}
                final_obj = Con(r_temp < 0, lb, r_temp)
                final_obj.save(corr_ras_path)
            else:
                r_temp.save(corr_ras_path)
        finally:
            r_temp = None
            try:
                arcpy.management.Delete(corr_ras_temp)
            except Exception:
                pass

        if clamp_info is not None:
            print(f"{akey}: min={clamp_info['min']} < 0，已将 <0 像元替换为下限 {clamp_info['lower_bound']}")

        eval_fc = os.path.join(gdb, f"eval_{uuid.uuid4().hex[:8]}")
        ExtractValuesToPoints(pts, arcpy.Raster(corr_ras_path), eval_fc, interpolate_values="INTERPOLATE", add_attributes="VALUE_ONLY")
        y_true = []
        y_pred = []
        with arcpy.da.SearchCursor(eval_fc, ["actual", "RASTERVALU"]) as cur:
            for a, p in cur:
                if a is None or p is None:
                    continue
                try:
                    aa = float(a)
                    pp = float(p)
                except Exception:
                    continue
                if math.isnan(pp) or math.isinf(pp):
                    continue
                y_true.append(aa)
                y_pred.append(pp)

        y_true_rf = [x[2] for x in rows]
        y_pred_rf = [x[3] for x in rows]
        r2_rf, rmse_rf = _r2_rmse(y_true_rf, y_pred_rf)
        r2_corr, rmse_corr = _r2_rmse(y_true, y_pred)

        accuracy_rows.append(
            {
                "attr": akey,
                "n": len(y_true),
                "r2_rf": r2_rf,
                "rmse_rf": rmse_rf,
                "r2_corr": r2_corr,
                "rmse_corr": rmse_corr,
            }
        )

        class_specs = standards.get(akey)
        class_disp = standards_display.get(akey)
        if class_specs and class_disp:
            rr = arcpy.Raster(corr_ras_path)
            base = SetNull(rr == rr, rr)
            outc = base
            for cls in range(1, 6):
                segs = class_specs.get(cls, [])
                cond = _segs_to_raster_cond(rr, segs)
                if cond is None:
                    continue
                outc = Con(cond, cls, outc)
            Int(outc).save(class_ras_path)

            map_stats = _calc_map_class_area_stats(class_ras_path)

            sidx = sample["header_idx"]
            values = []
            if akey in sidx:
                fi = sidx[akey]
                for r in sample["rows"]:
                    if fi < len(r):
                        values.append(_safe_float(r[fi]))
            sample_stats = _calc_sample_stats(values, class_specs)

            map_mm = _calc_raster_stats_with_median(corr_ras_path, gdb, seed=seed + 17, n_points=3000)
            sample_mm = sample_stats["mm"]

            _write_grade_sheet(wb_out, akey, class_disp, sample_stats, map_stats, sample_mm, map_mm)

        for p in [pts_gcs, eval_fc]:
            try:
                if arcpy.Exists(p):
                    arcpy.management.Delete(p)
            except Exception:
                pass
        if pts != pts_gcs:
            try:
                if arcpy.Exists(pts):
                    arcpy.management.Delete(pts)
            except Exception:
                pass

    accuracy_rows.sort(key=lambda x: x["attr"])
    out_accuracy_xlsx = os.path.join(out_tbl, "RF与叠加残差克里金_精度对比.xlsx")
    meta = {
        "rf_folder": rf_folder,
        "validation_xlsx": validation_xlsx,
        "sample_xlsx": sample_xlsx,
        "standard_xlsx": standard_xlsx,
        "output_root": out_root,
        "semivariogram": semivariogram,
        "neighbors": 12,
        "seed": seed,
    }
    _write_accuracy_xlsx(out_accuracy_xlsx, accuracy_rows, steps_text, meta)

    out_grade_xlsx = os.path.join(out_tbl, "叠加残差克里金_分级对比表.xlsx")
    if len(wb_out.sheetnames) == 0:
        wb_out.create_sheet("Empty")
    wb_out.save(out_grade_xlsx)

    print(out_root)


if __name__ == "__main__":
    main()
